#!/usr/bin/env python3
"""
Tool Spy Idea v1.0.0 — Desktop App
"""
import sys, os, subprocess

# PyInstaller support: detect bundled .exe
if getattr(sys, 'frozen', False):
    # Running as .exe (PyInstaller)
    APP_DIR_BASE = os.path.dirname(sys.executable)
    # Log errors to file vì --windowed ẩn console
    try:
        _log_path = os.path.join(APP_DIR_BASE, "app_log.txt")
        _log_f = open(_log_path, "w", encoding="utf-8")
        sys.stdout = _log_f
        sys.stderr = _log_f
    except: pass
else:
    APP_DIR_BASE = os.path.dirname(os.path.abspath(__file__))

def auto_install():
    if getattr(sys, 'frozen', False):
        return  # .exe đã có sẵn packages, không cần install
    required = {
        "flask": "flask",
        "playwright": "playwright",
        "bs4": "beautifulsoup4",
        "openpyxl": "openpyxl",
        "requests": "requests",
        "PIL": "Pillow",
        "dropbox": "dropbox",
    }
    missing = []
    for mod, pkg in required.items():
        try: __import__(mod)
        except ImportError: missing.append(pkg)
    if missing:
        print(f"📦 Cài: {', '.join(missing)}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing, "-q"])
        print("✓ Xong!")
    # Optional: pywebview (native window mode)
    try: __import__("webview")
    except ImportError:
        try: subprocess.check_call([sys.executable, "-m", "pip", "install", "pywebview", "-q"], stderr=subprocess.DEVNULL)
        except: pass  # Skip nếu không cài được (Python 3.14 chưa hỗ trợ)

auto_install()

import threading, json, webbrowser, time
from pathlib import Path

from flask import Flask, request, jsonify, send_from_directory, send_file

APP_DIR = Path(APP_DIR_BASE)

# PyInstaller 6.x puts --add-data files in _internal/
# Check both locations for static folder
_static_dir = APP_DIR / "static"
if not _static_dir.exists():
    _static_dir = APP_DIR / "_internal" / "static"
if not _static_dir.exists():
    _static_dir = APP_DIR / "static"  # fallback, will be created
    _static_dir.mkdir(exist_ok=True)

# Same for modules
_modules_dir = APP_DIR / "modules"
if not _modules_dir.exists() and (APP_DIR / "_internal" / "modules").exists():
    sys.path.insert(0, str(APP_DIR / "_internal"))

flask_app = Flask(__name__, static_folder=str(_static_dir))
sys.path.insert(0, str(APP_DIR))
DATA_DIR = APP_DIR / "data"
if not DATA_DIR.exists():
    DATA_DIR = APP_DIR / "_internal" / "data"
if not DATA_DIR.exists():
    DATA_DIR = APP_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DOWNLOADS_DIR = APP_DIR / "downloads"; DOWNLOADS_DIR.mkdir(exist_ok=True)
CONFIG_FILE = DATA_DIR / "config.json"
HISTORY_FILE = DATA_DIR / "spy_history.json"
APP_VERSION = "1.0.0"
# URL kiểm tra version mới (thay bằng URL của bạn, ví dụ: https://raw.githubusercontent.com/.../version.json)
UPDATE_CHECK_URL = "https://raw.githubusercontent.com/ken2112g/toolspyidea/refs/heads/main/version.json"  # để trống = không check update

def load_config():
    if CONFIG_FILE.exists():
        try: return json.loads(CONFIG_FILE.read_text("utf-8"))
        except: pass
    return {"browser_profile": None, "delay": 1.0, "proxy": "", "dropbox_token": "", "max_images": 10,
            "theme": "dark", "lang": "vi", "stats": {"spied":0,"downloaded":0,"cleaned":0,"dbx_links":0}}

def save_config(cfg):
    CONFIG_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), "utf-8")

# ===== Spy History helpers =====
def load_history():
    if HISTORY_FILE.exists():
        try: return json.loads(HISTORY_FILE.read_text("utf-8"))
        except: pass
    return []

def save_history(items):
    # Giới hạn 100 entries gần nhất để không bloat file
    HISTORY_FILE.write_text(json.dumps(items[-100:], indent=2, ensure_ascii=False), "utf-8")

def add_history_entry(entry):
    h = load_history()
    h.append(entry)
    save_history(h)

def bump_stat(key, n=1):
    """Tăng counter trong config.stats — fail-safe"""
    try:
        cfg = load_config()
        stats = cfg.get("stats", {})
        stats[key] = stats.get(key, 0) + n
        cfg["stats"] = stats
        save_config(cfg)
    except Exception: pass

# ===== URL Cache (24h TTL, LRU 500 entries) =====
URL_CACHE_FILE = DATA_DIR / "url_cache.json"
URL_CACHE_TTL = 24 * 3600  # 24 giờ
URL_CACHE_MAX = 500

def _cache_load():
    if URL_CACHE_FILE.exists():
        try: return json.loads(URL_CACHE_FILE.read_text("utf-8"))
        except: pass
    return {}

def _cache_save(cache):
    try: URL_CACHE_FILE.write_text(json.dumps(cache, ensure_ascii=False), "utf-8")
    except: pass

def cache_get(url):
    """Trả về data nếu cache còn hạn, hoặc None"""
    cache = _cache_load()
    entry = cache.get(url)
    if not entry: return None
    if time.time() - entry.get("ts", 0) > URL_CACHE_TTL:
        return None
    return entry.get("data")

def cache_set(url, data):
    """Lưu vào cache, LRU evict nếu vượt max"""
    cache = _cache_load()
    cache[url] = {"ts": int(time.time()), "data": data}
    # LRU evict
    if len(cache) > URL_CACHE_MAX:
        sorted_keys = sorted(cache.keys(), key=lambda k: cache[k].get("ts", 0))
        for k in sorted_keys[:len(cache) - URL_CACHE_MAX]:
            del cache[k]
    _cache_save(cache)

def cache_clear():
    _cache_save({})

# ===== Retry with exponential backoff =====
def retry_with_backoff(fn, max_tries=3, base_delay=2.0, *args, **kwargs):
    """Gọi fn(*args, **kwargs). Retry max_tries lần với delay 2s, 4s, 8s..."""
    import time as _t
    last_err = None
    for attempt in range(max_tries):
        try:
            return fn(*args, **kwargs)
        except Exception as e:
            last_err = e
            if attempt < max_tries - 1:
                _t.sleep(base_delay * (2 ** attempt))
    raise last_err

# ===== Audit log =====
AUDIT_LOG_FILE = DATA_DIR / "audit.log"
AUDIT_LOG_MAX_BYTES = 500_000  # ~500KB thì rotate

def audit_log(action, details=""):
    """Ghi log API call để debug sau"""
    try:
        # Rotate nếu quá lớn
        if AUDIT_LOG_FILE.exists() and AUDIT_LOG_FILE.stat().st_size > AUDIT_LOG_MAX_BYTES:
            AUDIT_LOG_FILE.rename(AUDIT_LOG_FILE.with_suffix(".log.old"))
        with AUDIT_LOG_FILE.open("a", encoding="utf-8") as f:
            ts = time.strftime('%Y-%m-%d %H:%M:%S')
            f.write(f"[{ts}] {action} {details}\n")
    except: pass

# ===== Chrome Extension State (cho Etsy) =====
_etsy_state = {"pending": [], "results": {}, "wanted_urls": set()}

@flask_app.route("/api/etsy/should-scrape")
def etsy_should_scrape():
    """Content script hỏi: URL này có được tool yêu cầu scrape không?"""
    import time as _t
    url = request.args.get("url", "")
    _etsy_state['_last_poll'] = _t.time()
    # Check match URL (với Etsy chỉ cần match listing ID)
    import re as _re
    m = _re.search(r'/listing/(\d+)', url)
    listing_id = m.group(1) if m else None
    
    wanted = False
    if listing_id:
        for want_url in _etsy_state["wanted_urls"]:
            if listing_id in want_url:
                wanted = True
                break
    
    print(f"[ETSY CHECK] URL={url[:80]}, listing={listing_id}, wanted={wanted}")
    resp = jsonify({"wanted": wanted})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@flask_app.route("/api/etsy/pending-urls")
def etsy_pending():
    """(Legacy) - vẫn dùng cho extension v1"""
    import time as _t
    _etsy_state['_last_poll'] = _t.time()
    urls = _etsy_state["pending"][:]
    _etsy_state["pending"] = []
    resp = jsonify({"urls": urls})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@flask_app.route("/api/etsy/receive", methods=["POST", "OPTIONS"])
def etsy_receive():
    """Extension gửi HTML về sau khi scrape xong"""
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        return resp
    d = request.json
    url = d.get("url", "")
    html = d.get("html", "")
    print(f"[ETSY RECEIVE] URL: {url[:80]}, HTML: {len(html)} chars")
    # Match với wanted URL qua listing ID
    import re as _re
    m = _re.search(r'/listing/(\d+)', url)
    listing_id = m.group(1) if m else None
    print(f"[ETSY RECEIVE] listing_id: {listing_id}, wanted: {list(_etsy_state['wanted_urls'])}")
    if listing_id:
        for want_url in list(_etsy_state["wanted_urls"]):
            if listing_id in want_url:
                _etsy_state["results"][want_url] = html
                _etsy_state["wanted_urls"].discard(want_url)
                print(f"[ETSY RECEIVE] ✓ Matched {want_url[:80]}")
                break
        else:
            print(f"[ETSY RECEIVE] No match for listing_id {listing_id}")
    resp = jsonify({"ok": True})
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

@flask_app.route("/api/etsy/extension-status")
def etsy_ext_status():
    """Check extension đã scrape Etsy gần đây không"""
    import time as _t
    last_poll = _etsy_state.get('_last_poll', 0)
    connected = last_poll > 0 and (_t.time() - last_poll) < 30
    return jsonify({"connected": connected})

@flask_app.route("/api/etsy/login", methods=["POST"])
def etsy_login():
    """Mở Chrome native với profile của tool để user login Etsy (không lag)"""
    import subprocess as _sp, platform as _plf
    
    tool_profile = str(APP_DIR / "etsy_profile")
    os.makedirs(tool_profile, exist_ok=True)
    
    # Tìm Chrome exe
    chrome_exe = None
    for cp in [
        os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
    ]:
        if os.path.exists(cp):
            chrome_exe = cp; break
    
    if not chrome_exe:
        return jsonify({"ok": False, "msg": "Không tìm thấy Chrome"})
    
    # Launch Chrome native với profile riêng — chạy bình thường, không lag
    try:
        _sp.Popen([
            chrome_exe,
            f"--user-data-dir={tool_profile}",
            "--no-first-run",
            "--no-default-browser-check",
            "https://www.etsy.com/",
        ])
        return jsonify({"ok": True, "msg": "Chrome đã mở, login xong đóng cửa sổ"})
    except Exception as e:
        return jsonify({"ok": False, "msg": f"Error: {e}"})

@flask_app.route("/api/etsy/check-session")
def etsy_check_session():
    """Kiểm tra xem profile đã có session Etsy hay chưa"""
    tool_profile = APP_DIR / "etsy_profile"
    
    if not tool_profile.exists():
        return jsonify({
            "has_profile": False,
            "has_cookies": False,
            "path": str(tool_profile),
            "message": "Profile chưa được tạo. Bấm 'Mở Chrome Login' để tạo."
        })
    
    # Tìm Cookies file ở bất kỳ vị trí nào trong profile folder
    # (Playwright có thể lưu ở Default/Cookies, hoặc path khác tùy version)
    cookies_info = []
    for f in tool_profile.rglob("Cookies"):
        if f.is_file():
            size = f.stat().st_size
            cookies_info.append({"path": str(f.relative_to(tool_profile)), "size": size})
    
    # Cũng check Local State (chứa session info)
    local_state = tool_profile / "Local State"
    
    # Session ok nếu có cookies file > 1000 bytes
    has_cookies = any(c["size"] > 1000 for c in cookies_info)
    
    return jsonify({
        "has_profile": True,
        "has_cookies": has_cookies,
        "cookies_files": cookies_info,
        "path": str(tool_profile),
        "message": f"Found {len(cookies_info)} cookies file(s)" if cookies_info else "Chưa tìm thấy cookies. Login lại để tạo session."
    })

@flask_app.route("/api/extension/open-folder", methods=["POST"])
def open_ext_folder():
    """Mở folder extension cho user load vào Chrome"""
    import subprocess, platform as plf
    ext_path = str(APP_DIR / "extension")
    try:
        if plf.system() == "Windows":
            subprocess.Popen(["explorer", ext_path])
        elif plf.system() == "Darwin":
            subprocess.Popen(["open", ext_path])
        else:
            subprocess.Popen(["xdg-open", ext_path])
        # Mở chrome://extensions
        webbrowser.open("chrome://extensions/")
        return jsonify({"ok": True, "path": ext_path})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "path": ext_path})

def _find_chrome_channel():
    """Detect Chrome channel for Playwright.
    Return None to use bundled browser first, fallback to system Chrome/Edge."""
    # Check if Playwright bundled browser exists
    try:
        import playwright
        from pathlib import Path
        browsers_dir = Path(playwright.__file__).parent / "driver" / "package" / ".local-browsers"
        if browsers_dir.exists() and any(browsers_dir.iterdir()):
            return None  # Dùng bundled browser
    except:
        pass
    
    # Fallback: dùng Chrome/Edge đã cài trên máy
    chrome_paths = [
        os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
    ]
    for p in chrome_paths:
        if os.path.exists(p):
            return "chrome"
    edge_paths = [
        os.path.expandvars(r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe"),
        os.path.expandvars(r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe"),
    ]
    for p in edge_paths:
        if os.path.exists(p):
            return "msedge"
    return None

def _launch_persistent(playwright, user_data_dir, headless=True, extra_args=None):
    """Launch persistent context, auto-detect channel"""
    channel = _find_chrome_channel()
    args = ["--disable-blink-features=AutomationControlled", "--no-first-run"]
    if extra_args:
        args.extend(extra_args)
    kwargs = {
        "user_data_dir": user_data_dir,
        "headless": headless,
        "args": args,
        "viewport": {"width": 1366, "height": 768},
    }
    if channel:
        kwargs["channel"] = channel
    return playwright.chromium.launch_persistent_context(**kwargs)

def launch_browser(playwright):
    """Launch Playwright browser cho non-Etsy sites"""
    import random
    from modules.scraper import STEALTH_JS, USER_AGENTS
    _cfg = load_config()
    _headless = _cfg.get("headless_scrape", True)
    _channel = _find_chrome_channel()
    launch_args = {"headless": _headless, "args": ["--disable-blink-features=AutomationControlled"]}
    if _channel:
        launch_args["channel"] = _channel
    browser = playwright.chromium.launch(**launch_args)
    ctx = browser.new_context(user_agent=random.choice(USER_AGENTS))
    page = ctx.new_page()
    page.add_init_script(STEALTH_JS)
    return ctx, page

def connect_real_chrome(playwright):
    """Kết nối vào Chrome thật đang chạy (cần Chrome đã bật debug port)"""
    import socket
    port = 9222
    # Kiểm tra Chrome có debug port không
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(2)
        s.connect(("127.0.0.1", port))
        s.close()
    except:
        return None, None, "Chrome chưa bật debug mode. Vào Settings → bấm 'Setup Chrome Debug Mode' trước."
    
    try:
        browser = playwright.chromium.connect_over_cdp(f"http://127.0.0.1:{port}")
        ctx = browser.contexts[0] if browser.contexts else browser.new_context()
        page = ctx.new_page()
        return ctx, page, None
    except Exception as e:
        return None, None, f"Không kết nối được: {e}"

@flask_app.route("/api/setup-chrome-debug", methods=["POST"])
def setup_chrome_debug():
    """Tạo shortcut Chrome với debug mode cho user"""
    import subprocess, glob
    
    # Tìm Chrome shortcut trên Desktop
    desktop = os.path.expandvars(r"%USERPROFILE%\Desktop")
    chrome_exe = None
    for p in [
        os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
        os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
    ]:
        if os.path.exists(p):
            chrome_exe = p; break
    
    if not chrome_exe:
        return jsonify({"ok": False, "msg": "Không tìm thấy Chrome!"})
    
    # Tạo file .bat trên Desktop
    bat_path = os.path.join(desktop, "Chrome Debug Mode.bat")
    bat_content = f'''@echo off
start "" "{chrome_exe}" --remote-debugging-port=9222
'''
    with open(bat_path, "w") as f:
        f.write(bat_content)
    
    return jsonify({"ok": True, "msg": f"Đã tạo shortcut trên Desktop: 'Chrome Debug Mode.bat'. Đóng Chrome → mở bằng shortcut đó → dùng tool bình thường.", "path": bat_path})

@flask_app.route("/api/check-chrome-debug")
def check_chrome_debug():
    """Kiểm tra Chrome debug mode đang chạy không"""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(2)
        s.connect(("127.0.0.1", 9222))
        s.close()
        return jsonify({"connected": True})
    except:
        return jsonify({"connected": False})

# =================== ROUTES ===================

@flask_app.route("/")
def index():
    return send_from_directory(str(APP_DIR / "static"), "index.html")

@flask_app.route("/favicon.ico")
def favicon():
    ico_path = APP_DIR / "static" / "favicon.ico"
    if not ico_path.exists():
        try:
            from PIL import Image, ImageDraw
            img = Image.new('RGBA', (64, 64), (0,0,0,0))
            d = ImageDraw.Draw(img)
            d.rounded_rectangle([1,1,63,63], radius=14, fill=(26,16,64))
            d.rounded_rectangle([1,1,63,63], radius=14, outline=(132,85,239), width=2)
            d.ellipse([17,17,47,47], outline=(186,158,255), width=2)
            d.line([32,8,32,22], fill=(83,221,252), width=2)
            d.line([32,42,32,56], fill=(83,221,252), width=2)
            d.line([8,32,22,32], fill=(83,221,252), width=2)
            d.line([42,32,56,32], fill=(83,221,252), width=2)
            d.ellipse([28,28,36,36], fill=(83,221,252))
            img.save(str(ico_path), format='ICO', sizes=[(64,64),(32,32),(16,16)])
        except: pass
    if ico_path.exists():
        return send_file(str(ico_path), mimetype='image/x-icon')
    return '', 204

@flask_app.route("/api/config", methods=["GET","POST"])
def api_config():
    if request.method == "POST":
        cfg = load_config(); cfg.update(request.json.get("config", {})); save_config(cfg)
        return jsonify({"status": "ok"})
    return jsonify(load_config())

# ===================== SPY HISTORY =====================
@flask_app.route("/api/history/list", methods=["GET"])
def history_list():
    """Trả về lịch sử spy (mới nhất trước)"""
    items = load_history()
    items.sort(key=lambda x: x.get("ts", 0), reverse=True)
    return jsonify({"items": items, "count": len(items)})

@flask_app.route("/api/history/add", methods=["POST"])
def history_add():
    """Thêm 1 entry lịch sử. Body: {type, label, count, data}"""
    d = request.json or {}
    entry = {
        "id": f"h_{int(time.time()*1000)}",
        "ts": int(time.time()),
        "type": d.get("type", "spy"),     # spy | shop | dropbox | download | clean
        "label": d.get("label", ""),        # hiển thị tên
        "count": d.get("count", 0),         # số item
        "data": d.get("data"),              # payload (có thể None để tiết kiệm dung lượng)
    }
    add_history_entry(entry)
    return jsonify({"status": "ok", "id": entry["id"]})

@flask_app.route("/api/history/delete/<hid>", methods=["POST"])
def history_delete(hid):
    items = [i for i in load_history() if i.get("id") != hid]
    save_history(items)
    return jsonify({"status": "ok"})

@flask_app.route("/api/history/clear", methods=["POST"])
def history_clear():
    save_history([])
    return jsonify({"status": "ok"})

# ===================== STATS =====================
@flask_app.route("/api/stats", methods=["GET"])
def api_stats():
    """Trả về stats counters"""
    cfg = load_config()
    return jsonify(cfg.get("stats", {"spied":0,"downloaded":0,"cleaned":0,"dbx_links":0}))

@flask_app.route("/api/stats/bump", methods=["POST"])
def api_stats_bump():
    """Bump counter. Body: {key: 'spied', n: 1}"""
    d = request.json or {}
    bump_stat(d.get("key", ""), int(d.get("n", 1)))
    cfg = load_config()
    return jsonify(cfg.get("stats", {}))

@flask_app.route("/api/stats/reset", methods=["POST"])
def api_stats_reset():
    cfg = load_config()
    cfg["stats"] = {"spied":0, "downloaded":0, "cleaned":0, "dbx_links":0}
    save_config(cfg)
    return jsonify(cfg["stats"])

# ===================== CACHE =====================
@flask_app.route("/api/cache/info", methods=["GET"])
def api_cache_info():
    """Info về URL cache"""
    cache = _cache_load()
    now = time.time()
    alive = sum(1 for e in cache.values() if now - e.get("ts", 0) <= URL_CACHE_TTL)
    return jsonify({"total": len(cache), "alive": alive, "ttl_hours": URL_CACHE_TTL/3600, "max": URL_CACHE_MAX})

@flask_app.route("/api/cache/clear", methods=["POST"])
def api_cache_clear():
    cache_clear()
    audit_log("CACHE_CLEARED")
    return jsonify({"ok": True})

# ===================== SEO SCORE (Clean Title) =====================
@flask_app.route("/api/clean/seo-score", methods=["POST"])
def api_clean_seo():
    """
    Tính SEO score cho list titles. Trả về score 0-100 + gợi ý.
    Factors:
    - Length (optimal 150-200 chars cho Amazon)
    - Keyword density
    - Has power words (gift/unique/custom/perfect)
    - Has benefit (for men/women/birthday/christmas)
    - No repeated words
    - No ALL CAPS
    """
    import re as _re
    d = request.json or {}
    titles = d.get("titles", [])
    
    POWER = ['gift', 'custom', 'personalized', 'unique', 'perfect', 'handmade', 'vintage',
             'premium', 'quality', 'original', 'exclusive', 'limited']
    BENEFIT = ['for men', 'for women', 'for kids', 'for him', 'for her',
               'birthday', 'christmas', 'wedding', 'anniversary', 'graduation',
               'mothers day', "mother's day", 'fathers day', "father's day",
               'valentine', 'halloween', 'easter']
    
    results = []
    for t in titles:
        score = 0
        tips = []
        length = len(t)
        words = t.split()
        words_lower = [w.lower() for w in words]
        t_lower = t.lower()
        
        # Length: 150-200 optimal = 30 points
        if 150 <= length <= 200:
            score += 30
        elif 100 <= length < 150:
            score += 20
            tips.append(f"Độ dài {length} — nên 150-200 ký tự")
        elif 80 <= length < 100:
            score += 10
            tips.append(f"Độ dài {length} — hơi ngắn")
        elif length > 200:
            score += 10
            tips.append(f"Độ dài {length} — Amazon cắt ở 200")
        else:
            tips.append(f"Quá ngắn ({length} ký tự)")
        
        # Power words = 20 points (5 per match, max 4 match)
        power_count = sum(1 for w in POWER if w in t_lower)
        score += min(power_count * 5, 20)
        if power_count == 0:
            tips.append("Thiếu power word (gift/custom/unique/perfect)")
        
        # Benefit/occasion = 20 points
        benefit_count = sum(1 for b in BENEFIT if b in t_lower)
        score += min(benefit_count * 10, 20)
        if benefit_count == 0:
            tips.append("Thiếu dịp/đối tượng (for men/birthday/christmas)")
        
        # Không có duplicate word = 15 points
        from collections import Counter
        cnt = Counter(w for w in words_lower if len(w) > 3)
        max_rep = max(cnt.values()) if cnt else 0
        if max_rep <= 1:
            score += 15
        elif max_rep == 2:
            score += 10
        else:
            score += 0
            tips.append(f"Có từ lặp {max_rep} lần")
        
        # Không ALL CAPS = 10 points
        all_caps_words = [w for w in words if len(w) > 2 and w.isupper()]
        if len(all_caps_words) <= 1:
            score += 10
        elif len(all_caps_words) <= 3:
            score += 5
        else:
            tips.append(f"Quá nhiều từ viết HOA ({len(all_caps_words)})")
        
        # Có số = 5 points (giúp specific)
        if any(ch.isdigit() for ch in t):
            score += 5
        
        # Grade
        if score >= 80: grade = "A"
        elif score >= 65: grade = "B"
        elif score >= 50: grade = "C"
        else: grade = "D"
        
        results.append({
            "title": t,
            "score": score,
            "grade": grade,
            "length": length,
            "power_words": power_count,
            "benefits": benefit_count,
            "max_repeat": max_rep,
            "tips": tips,
        })
    return jsonify({"results": results})

# ===================== VERSION / UPDATE CHECK =====================
@flask_app.route("/api/version", methods=["GET"])
def api_version():
    """Trả version hiện tại + check update nếu có URL"""
    result = {"current": APP_VERSION, "latest": None, "has_update": False, "message": ""}
    if not UPDATE_CHECK_URL:
        result["message"] = "Update check không được cấu hình"
        return jsonify(result)
    try:
        import urllib.request, json as _json
        with urllib.request.urlopen(UPDATE_CHECK_URL, timeout=4) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
            latest = str(data.get("version", "")).strip()
            result["latest"] = latest
            result["notes"] = data.get("notes", "")
            result["download_url"] = data.get("download_url", "")
            # So sánh version đơn giản: "1.0.0" < "1.1.0"
            def _v(s): return tuple(int(x) for x in s.split(".") if x.isdigit())
            if latest and _v(latest) > _v(APP_VERSION):
                result["has_update"] = True
                result["message"] = f"Có bản mới: v{latest}"
            else:
                result["message"] = "Bạn đang dùng bản mới nhất"
    except Exception as e:
        result["message"] = f"Không check được: {e}"
    return jsonify(result)

# ===================== CSV EXPORTS =====================
def _csv_response(rows, filename):
    """Tạo CSV response từ list of lists (hàng đầu = header)"""
    import csv as _csv, io as _io
    buf = _io.StringIO()
    buf.write('\ufeff')  # BOM cho Excel nhận UTF-8
    w = _csv.writer(buf, quoting=_csv.QUOTE_MINIMAL)
    for row in rows:
        w.writerow(row)
    data = buf.getvalue().encode("utf-8")
    bio = _io.BytesIO(data); bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="text/csv; charset=utf-8")

@flask_app.route("/api/spy/export-csv", methods=["POST"])
def spy_export_csv():
    d = request.json or {}
    results = d.get("results", [])
    max_images = max((len(r.get("images", [])) for r in results), default=0)
    headers = ["URL", "Title"] + [f"Image {i+1}" for i in range(max_images)]
    rows = [headers]
    for r in results:
        row = [r.get("url",""), r.get("title","")] + list(r.get("images", []))
        while len(row) < 2 + max_images: row.append("")
        rows.append(row)
    return _csv_response(rows, "spy_export.csv")

@flask_app.route("/api/spy/etsy-shop-export-csv", methods=["POST"])
def etsy_shop_export_csv():
    d = request.json or {}
    sections = d.get("sections", [])
    # Layout: cột = section, rows = URLs
    max_rows = max((len(s.get('products') or s.get('links') or []) for s in sections), default=0)
    headers = [s['name'] for s in sections]
    rows = [headers]
    for ri in range(max_rows):
        row = []
        for sec in sections:
            items = sec.get('products') or sec.get('links') or []
            if ri < len(items):
                it = items[ri]
                row.append(it if isinstance(it, str) else it.get('url',''))
            else:
                row.append("")
        rows.append(row)
    return _csv_response(rows, "shop_sections.csv")

@flask_app.route("/api/clean/export-csv", methods=["POST"])
def clean_export_csv():
    d = request.json or {}
    results = d.get("results", [])
    headers = ["#", "Original", "Processed", "Length", "Copyright Removed", "Duplicates Fixed", "Warnings"]
    rows = [headers]
    for i, r in enumerate(results, 1):
        removed = ", ".join(r.get("removed_keywords", []))
        dedup = " | ".join(f'{c["original"]}→{c["new"]}' for c in r.get("dedup_changes", []))
        warns = " | ".join(w["message"] for w in r.get("warnings", []))
        rows.append([i, r.get("original",""), r.get("fixed",""), r.get("length",0), removed, dedup, warns])
    return _csv_response(rows, "clean_titles.csv")

@flask_app.route("/api/dropbox/export-csv", methods=["POST"])
def dbx_export_csv():
    d = request.json or {}
    results = d.get("results", {})  # {folder_name: [link1, link2, ...]}
    # Chuẩn hóa input — có thể là dict hoặc list of {name, links}
    if isinstance(results, list):
        folders = [(r.get("name") or r.get("folder") or "", r.get("links") or []) for r in results]
    else:
        folders = list(results.items())
    max_links = max((len(links) for _, links in folders), default=0)
    headers = ["Folder"] + [f"Link {i+1}" for i in range(max_links)]
    rows = [headers]
    for name, links in folders:
        row = [name] + list(links)
        while len(row) < 1 + max_links: row.append("")
        rows.append(row)
    return _csv_response(rows, "dropbox_links.csv")

@flask_app.route("/api/browser-profiles")
def get_profiles():
    from modules.browser_detect import detect_all_profiles
    return jsonify({"profiles": detect_all_profiles()})

@flask_app.route("/api/keywords")
def get_keywords():
    from modules.title_cleaner import load_keywords
    kws = load_keywords(); return jsonify({"keywords": kws, "count": len(kws)})

@flask_app.route("/api/keywords/add", methods=["POST"])
def api_add_kw():
    from modules.title_cleaner import add_keyword, load_keywords
    add_keyword(request.json["keyword"]); kws = load_keywords()
    return jsonify({"keywords": kws, "count": len(kws)})

@flask_app.route("/api/keywords/remove", methods=["POST"])
def api_remove_kw():
    from modules.title_cleaner import remove_keyword, load_keywords
    remove_keyword(request.json["keyword"]); kws = load_keywords()
    return jsonify({"keywords": kws, "count": len(kws)})

@flask_app.route("/api/clean/process", methods=["POST"])
def clean_process():
    from modules.title_cleaner import process_titles, load_keywords
    d = request.json
    results = process_titles(
        d["titles"],
        load_keywords(),
        enable_clean=d.get("enable_clean", True),
        enable_dedup=d.get("enable_dedup", True),
        whole_word=d.get("whole_word", True),
        case_sensitive=d.get("case_sensitive", False),
        max_repeat=d.get("max_repeat", 2)
    )
    return jsonify({"results": results})

@flask_app.route("/api/clean/import-titles", methods=["POST"])
def clean_import_titles():
    """Import titles từ Excel/CSV/TXT — scan mọi cột lấy text"""
    try:
        f = request.files.get('file')
        if not f: return jsonify({"error": "No file"}), 400
        import tempfile, os as _os
        ext = _os.path.splitext(f.filename)[1].lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            f.save(tmp.name); tmp_path = tmp.name
        titles = []
        try:
            if ext in ('.txt',):
                with open(tmp_path, encoding='utf-8', errors='ignore') as fp:
                    titles = [ln.strip() for ln in fp if ln.strip()]
            elif ext == '.csv':
                import csv
                with open(tmp_path, encoding='utf-8-sig') as cf:
                    for row in csv.reader(cf):
                        for cell in row:
                            cs = str(cell).strip()
                            if cs and len(cs) > 5 and not cs.startswith('http'):
                                titles.append(cs)
            else:
                from openpyxl import load_workbook
                wb = load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value:
                            cs = str(cell.value).strip()
                            if len(cs) > 5 and not cs.startswith('http'):
                                titles.append(cs)
        finally:
            try: _os.unlink(tmp_path)
            except: pass
        # Dedupe giữ thứ tự
        seen = set(); unique = []
        for t in titles:
            if t not in seen: seen.add(t); unique.append(t)
        return jsonify({"titles": unique})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@flask_app.route("/api/clean/export", methods=["POST"])
def clean_export():
    """Export results ra Excel: Original | Processed | Removed | Fixes | Warnings"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io
    d = request.json
    results = d.get("results", [])
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Titles"
    ws.append(["#", "Original", "Processed", "Length", "Copyright Removed", "Duplicates Fixed", "Warnings"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8455EF")
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(results, 1):
        removed = ", ".join(r.get("removed_keywords", []))
        dedup = " | ".join(f'{c["original"]}→{c["new"]}' for c in r.get("dedup_changes", []))
        warns = " | ".join(w["message"] for w in r.get("warnings", []))
        ws.append([i, r.get("original",""), r.get("fixed",""), r.get("length",0), removed, dedup, warns])
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 40
    # Wrap text để không tràn ô
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"clean_titles.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@flask_app.route("/api/download/sku-autofill", methods=["POST"])
def sku_autofill():
    from modules.downloader import auto_fill_sku
    d = request.json; return jsonify({"skus": auto_fill_sku(d["first_sku"], d["count"])})

@flask_app.route("/api/spy/import-urls", methods=["POST"])
def spy_import_urls():
    """Import URLs từ file Excel/CSV - chấp nhận mọi cột có URL"""
    try:
        f = request.files.get('file')
        if not f: return jsonify({"error": "No file"}), 400
        
        import tempfile, os as _os
        ext = _os.path.splitext(f.filename)[1].lower()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name
        
        urls = []
        try:
            if ext == '.csv':
                import csv
                with open(tmp_path, encoding='utf-8-sig') as cf:
                    for row in csv.reader(cf):
                        for cell in row:
                            cs = str(cell).strip().strip('"\'')
                            if cs.startswith('http'): urls.append(cs)
            else:
                from openpyxl import load_workbook
                wb = load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value:
                            cs = str(cell.value).strip().strip('"\'')
                            if cs.startswith('http'): urls.append(cs)
        finally:
            try: _os.unlink(tmp_path)
            except: pass
        
        # Dedupe giữ thứ tự
        seen = set(); unique = []
        for u in urls:
            if u not in seen:
                seen.add(u); unique.append(u)
        
        return jsonify({"urls": unique})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@flask_app.route("/api/download/pick-folder", methods=["POST"])
def download_pick_folder():
    """Mở native folder picker dialog và return path chọn"""
    try:
        import tkinter as tk
        from tkinter import filedialog
        import threading
        result = {"path": ""}
        
        def show_dialog():
            root = tk.Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            path = filedialog.askdirectory(title="Chọn folder lưu ảnh", parent=root)
            result["path"] = path or ""
            root.destroy()
        
        # Chạy dialog trên thread chính (tkinter cần main thread)
        t = threading.Thread(target=show_dialog)
        t.start()
        t.join(timeout=120)  # tối đa 2 phút
        
        return jsonify({"path": result["path"]})
    except Exception as e:
        return jsonify({"path": "", "error": str(e)})

@flask_app.route("/api/download/import-excel", methods=["POST"])
def import_excel():
    """Import Excel/CSV với format: Col A = SKU, Col B+ = image URLs"""
    try:
        f = request.files.get('file')
        if not f: return jsonify({"error": "No file"}), 400
        
        import tempfile, os as _os
        ext = _os.path.splitext(f.filename)[1].lower()
        
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name
        
        products = []
        try:
            if ext == '.csv':
                import csv
                with open(tmp_path, encoding='utf-8-sig') as cf:
                    reader = csv.reader(cf)
                    for row in reader:
                        row = [c for c in row if c and c.strip()]
                        if not row: continue
                        sku = row[0].strip()
                        if not sku or sku.lower() in ('sku', 'title'): continue
                        imgs = [c.strip() for c in row[1:] if c.strip().startswith('http')]
                        if imgs:
                            products.append({"sku": sku, "title": sku, "images": imgs})
            else:
                from openpyxl import load_workbook
                wb = load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb.active
                for row_idx, row in enumerate(ws.iter_rows(min_row=1), 1):
                    cells = [c.value for c in row if c.value is not None]
                    if not cells: continue
                    sku = str(cells[0]).strip()
                    if not sku: continue
                    # Skip header row
                    if row_idx == 1 and sku.lower() in ('sku', 'title', 'name'): continue
                    imgs = []
                    title = ''
                    for val in cells[1:]:
                        vs = str(val).strip()
                        if vs.startswith('http'):
                            imgs.append(vs)
                        elif not title and len(vs) > 5:
                            title = vs
                    if imgs:
                        products.append({"sku": sku, "title": title or sku, "images": imgs})
        finally:
            try: _os.unlink(tmp_path)
            except: pass
        
        return jsonify({"products": products})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# =================== DOWNLOAD JOBS ===================
_download_jobs = {}  # job_id -> {status, progress, logs, total, done, output_dir}

def _run_download_job(job_id, products, delay, custom_output=None):
    """Tải ảnh trong thread với streaming progress + parallel per-product"""
    from modules.downloader import download_image
    from pathlib import Path
    from datetime import datetime
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import time, re as _re
    
    job = _download_jobs[job_id]
    def log(m, t='info'):
        job["logs"].append({"m": m, "t": t, "ts": time.strftime('%H:%M:%S')})
    
    try:
        # Output folder - dùng custom path nếu có, fallback Downloads
        stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        if custom_output:
            base = Path(custom_output)
            base.mkdir(parents=True, exist_ok=True)
            output_dir = base / f"spy_{stamp}"
        else:
            home = Path.home()
            base = home / "Downloads" / "toolspyidea-downloads"
            base.mkdir(parents=True, exist_ok=True)
            output_dir = base / stamp
        output_dir.mkdir(exist_ok=True)
        job["output_dir"] = str(output_dir)
        log(f"Output folder: {output_dir}")
        
        total_items = sum(len(p.get('images', [])) * max(1, p.get('qty', 1)) for p in products)
        job["total"] = total_items
        done = 0
        titles_rows = []  # (SKU, Title, URL) để xuất Excel cuối
        
        for pi, prod in enumerate(products):
            sku = prod.get('sku', '').strip() or f"product_{pi+1}"
            images = prod.get('images', [])
            qty = max(1, int(prod.get('qty', 1)))
            title = prod.get('title', '')
            prod_url = prod.get('url', '')
            
            if not images:
                log(f"[{pi+1}/{len(products)}] {sku}: no images, skip", 'warn')
                continue
            
            log(f"[{pi+1}/{len(products)}] {sku} — {len(images)} images × qty {qty}")
            
            # Qty > 1 → tạo nhiều folder với SKU tuần tự
            skus_to_create = [sku]
            if qty > 1:
                m = _re.search(r'(\d+)$', sku)
                if m:
                    num_str = m.group(1)
                    prefix = sku[:m.start()]
                    num_len = len(num_str)
                    start = int(num_str)
                    skus_to_create = [f"{prefix}{str(start+i).zfill(num_len)}" for i in range(qty)]
                else:
                    skus_to_create = [f"{sku}-{i+1}" for i in range(qty)]
            
            for cur_sku in skus_to_create:
                sku_dir = output_dir / cur_sku
                sku_dir.mkdir(exist_ok=True)
                titles_rows.append((cur_sku, title, prod_url))
                
                # Lấy template từ config — default '{sku}_{i}'
                from modules.downloader import apply_filename_template
                tmpl = load_config().get("filename_template", "{sku}_{i}")
                
                # Parallel download 8 ảnh cùng lúc
                def _dl_one(idx, url):
                    filename = apply_filename_template(tmpl, cur_sku, idx, len(images), url)
                    filepath = str(sku_dir / filename)
                    try:
                        return (idx, filename, download_image(url, filepath), None)
                    except Exception as e:
                        return (idx, filename, False, str(e))
                
                with ThreadPoolExecutor(max_workers=8) as executor:
                    futures = [executor.submit(_dl_one, ii, img_url) for ii, img_url in enumerate(images, 1)]
                    for future in as_completed(futures):
                        idx, filename, success, err = future.result()
                        if success:
                            job["success"] = job.get("success", 0) + 1
                        else:
                            job["failed"] = job.get("failed", 0) + 1
                            log(f"  ✗ {filename}: {err or 'failed'}", 'err')
                        done += 1
                        job["done"] = done
        
        # Xuất Excel titles.xlsx ở root folder
        if titles_rows:
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Titles"
                ws.append(["SKU", "Title", "Source URL"])
                # Style header
                from openpyxl.styles import Font, PatternFill, Alignment
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="8455EF")
                # Data
                for row in titles_rows:
                    ws.append(list(row))
                # Auto width
                ws.column_dimensions['A'].width = 28
                ws.column_dimensions['B'].width = 80
                ws.column_dimensions['C'].width = 60
                # Wrap text để không tràn ô khi copy sang Google Sheets
                wrap = Alignment(wrap_text=True, vertical="top")
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = wrap
                xlsx_path = output_dir / "titles.xlsx"
                wb.save(xlsx_path)
                log(f"📄 titles.xlsx saved ({len(titles_rows)} rows)", 'ok')
            except Exception as xe:
                log(f"Excel export error: {xe}", 'warn')
        
        log(f"✓ Done. {job.get('success',0)} saved, {job.get('failed',0)} failed.", 'ok')
        log(f"📁 Folder: {output_dir}", 'ok')
    except Exception as e:
        import traceback
        log(f"FATAL: {e}", 'err')
        traceback.print_exc()
    
    job["status"] = "done"

@flask_app.route("/api/download/batch-start", methods=["POST"])
def download_batch_start():
    import threading as _th, uuid
    d = request.json
    products = d.get("products", [])
    delay = d.get("delay", 0.3)
    custom_output = d.get("output_dir", "").strip() or None
    
    job_id = uuid.uuid4().hex[:12]
    _download_jobs[job_id] = {
        "status": "running",
        "logs": [],
        "total": 0,
        "done": 0,
        "success": 0,
        "failed": 0,
        "output_dir": ""
    }
    t = _th.Thread(target=_run_download_job, args=(job_id, products, delay, custom_output), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})

@flask_app.route("/api/download/batch-status/<job_id>")
def download_batch_status(job_id):
    job = _download_jobs.get(job_id)
    if not job: return jsonify({"error": "Job not found"}), 404
    return jsonify(job)

@flask_app.route("/api/download/open-folder", methods=["POST"])
def download_open_folder():
    """Mở folder download trong Explorer"""
    import subprocess, platform
    d = request.json
    path = d.get("path", "")
    if not path: return jsonify({"ok": False, "error": "No path"}), 400
    try:
        if platform.system() == "Windows":
            subprocess.Popen(["explorer", path])
        elif platform.system() == "Darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# =================== SPY JOBS (streaming) ===================
_spy_jobs = {}  # job_id -> {status, results, logs, total, done}

def _cleanup_old_jobs():
    """Auto-clean jobs older than 1 hour"""
    import time as _t
    now = _t.time()
    for store in [_spy_jobs, _download_jobs]:
        to_del = [k for k, v in store.items() if v.get("_created", 0) < now - 3600 and v.get("status") in ("done", "stopped")]
        for k in to_del:
            del store[k]

def _run_spy_job(job_id, urls, max_img):
    """Chạy spy trong thread, update _spy_jobs[job_id] liên tục"""
    from modules.scraper import extract_product_data, detect_platform, EtsyExtractor, ShopifyExtractor, STEALTH_JS, USER_AGENTS
    import random, time, traceback, re as _re
    from playwright.sync_api import sync_playwright
    
    job = _spy_jobs[job_id]
    cfg = load_config()
    
    def log(m, t='info'):
        job["logs"].append({"m": m, "t": t, "ts": time.strftime('%H:%M:%S')})
    
    def push(data):
        job["results"].append(data)
        job["done"] = len(job["results"])
    
    def check_signal():
        """Return True if should continue, False if stopped. Blocks while paused."""
        while job.get("signal") == "pause":
            time.sleep(0.5)
        return job.get("signal") != "stop"
    
    # Normalize URLs
    clean_urls = []
    for u in urls:
        u = u.strip()
        if 'etsy.com' in u and '/listing/' in u:
            norm = EtsyExtractor.normalize_link(u)
            clean_urls.append(norm if norm else u)
        else:
            clean_urls.append(u)
    
    job["total"] = len(clean_urls)
    log(f"INITIALIZING SPY ENGINE... ({len(clean_urls)} URLs)")
    
    try:
        with sync_playwright() as p:
            delay = cfg.get("delay", 1.0)
            etsy_urls = [u for u in clean_urls if 'etsy.com' in u]
            other_urls = [u for u in clean_urls if 'etsy.com' not in u]
            
            # ===== NON-ETSY =====
            if other_urls:
                # Phân loại: Shopify (dùng HTTP JSON API, parallel) + Others (cần browser)
                shopify_urls = []
                generic_urls = []
                for url in other_urls:
                    m = _re.search(r'/products/([^/?#]+)', url)
                    if m and detect_platform(url) not in ('etsy', 'amazon'):
                        shopify_urls.append(url)
                    else:
                        generic_urls.append(url)
                
                # === Shopify: HTTP parallel (5 threads, nhanh gấp 5-10 lần) ===
                if shopify_urls:
                    from concurrent.futures import ThreadPoolExecutor, as_completed
                    import requests as _rq
                    from urllib.parse import urlparse as _urlparse
                    from modules.scraper import _clean_title
                    
                    def _fetch_shopify_once(url):
                        m = _re.search(r'/products/([^/?#]+)', url)
                        handle = m.group(1)
                        parsed = _urlparse(url)
                        json_url = f"{parsed.scheme}://{parsed.netloc}/products/{handle}.json"
                        r = _rq.get(json_url, timeout=15, headers={
                            "User-Agent": random.choice(USER_AGENTS)
                        })
                        r.raise_for_status()
                        jdata = r.json()
                        product = jdata.get("product", {})
                        if not product:
                            raise ValueError("empty product")
                        title = _clean_title(product.get("title", "(No Title)"))
                        imgs = [img.get("src", "") for img in product.get("images", [])]
                        imgs = [i for i in imgs if i][:max_img]
                        return {"url": url, "title": title, "images": imgs, "platform": detect_platform(url)}
                    
                    def _fetch_shopify(url):
                        # CACHE first
                        cached = cache_get(url)
                        if cached:
                            audit_log("SPY_CACHE_HIT", url)
                            return (url, cached, None)
                        try:
                            # Retry 3 lần với backoff 2s, 4s, 8s
                            data = retry_with_backoff(_fetch_shopify_once, 3, 2.0, url)
                            cache_set(url, data)
                            audit_log("SPY_SHOPIFY_OK", f"{url} | {len(data.get('images',[]))} imgs")
                            return (url, data, None)
                        except Exception as e:
                            audit_log("SPY_SHOPIFY_FAIL", f"{url} | {e}")
                            return (url, None, str(e))
                    
                    log(f"Shopify: parallel fetching {len(shopify_urls)} URLs...")
                    with ThreadPoolExecutor(max_workers=5) as ex:
                        futures = [ex.submit(_fetch_shopify, u) for u in shopify_urls]
                        for future in as_completed(futures):
                            url, data, err = future.result()
                            if data:
                                push(data)
                                log(f"✓ {data['title'][:40]} | {len(data['images'])} images", 'ok')
                            else:
                                # Fallback: thêm vào generic để thử browser
                                generic_urls.append(url)
                                log(f"Shopify JSON fail {url[:50]}: {err} — fallback browser", 'warn')
                
                # === Generic: browser sequential ===
                if generic_urls:
                    # Check cache first → tách URL đã cache ra
                    browser_urls = []
                    for url in generic_urls:
                        cached = cache_get(url)
                        if cached:
                            push(cached)
                            log(f"[cache] {cached['title'][:40]} | {len(cached.get('images',[]))} images", 'info')
                            audit_log("SPY_CACHE_HIT", url)
                        else:
                            browser_urls.append(url)
                    if browser_urls:
                        ctx, page = launch_browser(p)
                        for url in browser_urls:
                            if not check_signal(): break
                            try:
                                platform = detect_platform(url)
                                log(f"{platform}: {url[:60]}...")
                                page.goto(url, timeout=60000, wait_until="domcontentloaded")
                                time.sleep(delay)
                                page.mouse.wheel(0, 3000); time.sleep(1)
                                page.mouse.wheel(0, 3000); time.sleep(1)
                                html = page.content()
                                data = extract_product_data(html, url, max_img)
                                push(data)
                                cache_set(url, data)
                                audit_log("SPY_GENERIC_OK", f"{url} | {len(data.get('images',[]))} imgs")
                                log(f"✓ {data['title'][:40]} | {len(data.get('images',[]))} images", 'ok')
                            except Exception as e:
                                push({"url": url, "title": f"Error: {str(e)[:80]}", "images": [], "platform": detect_platform(url), "error": True})
                                audit_log("SPY_GENERIC_FAIL", f"{url} | {e}")
                                log(f"✗ {url[:60]}: {e}", 'err')
                        try: ctx.close()
                        except: pass
            
            # ===== ETSY (persistent profile) =====
            if etsy_urls:
                # Check cache first
                cached_etsy = []
                fresh_etsy = []
                for url in etsy_urls:
                    c = cache_get(url)
                    if c:
                        cached_etsy.append((url, c))
                    else:
                        fresh_etsy.append(url)
                for url, data in cached_etsy:
                    push(data)
                    log(f"[cache] Etsy {data['title'][:40]} | {len(data.get('images',[]))} images", 'info')
                    audit_log("SPY_CACHE_HIT", url)
                
                if fresh_etsy:
                    log(f"Etsy: loading {len(fresh_etsy)} URLs via tool profile...")
                    tool_profile = str(APP_DIR / "etsy_profile")
                    os.makedirs(tool_profile, exist_ok=True)
                    _cfg = load_config()
                    _headless = _cfg.get("headless_scrape", True)
                    try:
                        ctx_etsy = _launch_persistent(p, tool_profile, headless=_headless)
                        page_etsy = ctx_etsy.pages[0] if ctx_etsy.pages else ctx_etsy.new_page()
                        page_etsy.add_init_script(STEALTH_JS)
                        
                        # Hi-res option từ config (user có thể toggle trong Settings)
                        hi_res = _cfg.get("etsy_hi_res", False)
                        
                        for url in fresh_etsy:
                            if not check_signal(): break
                            try:
                                log(f"Etsy: {url[:60]}...")
                                page_etsy.goto(url, timeout=30000, wait_until="domcontentloaded")
                                time.sleep(3)
                                html = page_etsy.content()
                                if len(html) < 10000:
                                    time.sleep(5)
                                    html = page_etsy.content()
                                if len(html) > 10000:
                                    page_etsy.mouse.wheel(0, 2000); time.sleep(1)
                                    page_etsy.mouse.wheel(0, 2000); time.sleep(1)
                                    html = page_etsy.content()
                                    data = extract_product_data(html, url, max_img, hi_res=hi_res)
                                    push(data)
                                    cache_set(url, data)
                                    audit_log("SPY_ETSY_OK", f"{url} | {len(data.get('images',[]))} imgs | hi_res={hi_res}")
                                    log(f"✓ {data['title'][:40]} | {len(data.get('images',[]))} images", 'ok')
                                else:
                                    push({"url": url, "title": "Etsy blocked - login lại qua Settings", "images": [], "platform": "etsy"})
                                    audit_log("SPY_ETSY_BLOCKED", url)
                                    log(f"✗ Etsy blocked: {url[:60]}", 'err')
                            except Exception as e:
                                push({"url": url, "title": f"Error: {str(e)[:80]}", "images": [], "platform": "etsy"})
                                audit_log("SPY_ETSY_FAIL", f"{url} | {e}")
                                log(f"✗ {e}", 'err')
                        ctx_etsy.close()
                    except Exception as e:
                        log(f"Etsy profile error: {e}", 'err')
                        for url in fresh_etsy:
                            push({"url": url, "title": f"Profile error: {str(e)[:80]}", "images": [], "platform": "etsy"})
    except Exception as e:
        log(f"FATAL: {e}", 'err')
        traceback.print_exc()
    
    if job.get("signal") == "stop":
        job["status"] = "stopped"
        log(f"STOPPED: {len(job['results'])} products collected.", 'warn')
    else:
        job["status"] = "done"
        log(f"COMPLETE: {len(job['results'])} products.", 'ok')

@flask_app.route("/api/spy/start", methods=["POST"])
def spy_start():
    import threading as _th, uuid
    d = request.json
    urls = d.get("urls", [])
    max_img = d.get("max_images", 10)
    job_id = uuid.uuid4().hex[:12]
    _spy_jobs[job_id] = {"status": "running", "results": [], "logs": [], "total": len(urls), "done": 0, "signal": "run", "_created": time.time()}
    t = _th.Thread(target=_run_spy_job, args=(job_id, urls, max_img), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})

@flask_app.route("/api/spy/pause/<job_id>", methods=["POST"])
def spy_pause(job_id):
    job = _spy_jobs.get(job_id)
    if not job: return jsonify({"error": "Not found"}), 404
    job["signal"] = "pause"
    job["status"] = "paused"
    return jsonify({"ok": True})

@flask_app.route("/api/spy/resume/<job_id>", methods=["POST"])
def spy_resume(job_id):
    job = _spy_jobs.get(job_id)
    if not job: return jsonify({"error": "Not found"}), 404
    job["signal"] = "run"
    job["status"] = "running"
    return jsonify({"ok": True})

@flask_app.route("/api/spy/stop/<job_id>", methods=["POST"])
def spy_stop(job_id):
    job = _spy_jobs.get(job_id)
    if not job: return jsonify({"error": "Not found"}), 404
    job["signal"] = "stop"
    job["status"] = "stopped"
    return jsonify({"ok": True})

@flask_app.route("/api/spy/retry", methods=["POST"])
def spy_retry():
    """Retry 1 hoặc nhiều URL bị lỗi"""
    import threading as _th, uuid
    d = request.json
    urls = d.get("urls", [])
    max_img = d.get("max_images", 10)
    if not urls: return jsonify({"error": "No URLs"}), 400
    job_id = uuid.uuid4().hex[:12]
    _spy_jobs[job_id] = {"status": "running", "results": [], "logs": [], "total": len(urls), "done": 0, "signal": "run", "_created": time.time()}
    t = _th.Thread(target=_run_spy_job, args=(job_id, urls, max_img), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})

@flask_app.route("/api/spy/status/<job_id>")
def spy_status(job_id):
    job = _spy_jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job)

@flask_app.route("/api/spy/export-excel", methods=["POST"])
def spy_export_excel():
    """Export results ra Excel: col 1=URL, col 2=Title, col 3+=mỗi ảnh 1 cột"""
    from openpyxl import Workbook
    import io as _io
    d = request.json
    results = d.get("results", [])
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Spy Results"
    
    # Tìm max số images để biết bao nhiêu cột
    max_images = max((len(r.get("images", [])) for r in results), default=0)
    
    # Header
    headers = ["URL", "Title"] + [f"Image {i+1}" for i in range(max_images)]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8455EF")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for r in results:
        row = [r.get("url", ""), r.get("title", "")] + list(r.get("images", []))
        while len(row) < 2 + max_images:
            row.append("")
        ws.append(row)
    
    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 60
    for i in range(max_images):
        col_letter = chr(ord('C') + i) if i < 24 else 'A' + chr(ord('A') + i - 24)
        ws.column_dimensions[col_letter].width = 55
    
    # Wrap text để không tràn ô khi copy sang Google Sheets
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="spy_export.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@flask_app.route("/api/spy/products", methods=["POST"])
def spy_products():
    from modules.scraper import extract_product_data, detect_platform, EtsyExtractor, ShopifyExtractor, STEALTH_JS, USER_AGENTS
    import random, time, traceback, re as _re
    from playwright.sync_api import sync_playwright
    d = request.json; urls = d.get("urls",[]); max_img = d.get("max_images",10)
    cfg = load_config(); results = []
    
    # Normalize URLs
    clean_urls = []
    for u in urls:
        u = u.strip()
        if 'etsy.com' in u and '/listing/' in u:
            norm = EtsyExtractor.normalize_link(u)
            clean_urls.append(norm if norm else u)
        else:
            clean_urls.append(u)
    
    try:
        with sync_playwright() as p:
            delay = cfg.get("delay",1.0)
            
            # Tách Etsy và non-Etsy URLs
            etsy_urls = [u for u in clean_urls if 'etsy.com' in u]
            other_urls = [u for u in clean_urls if 'etsy.com' not in u]
            
            # ===== XỬ LÝ NON-ETSY (Chrome Playwright) =====
            if other_urls:
                ctx, page = launch_browser(p)
                for url in other_urls:
                    try:
                        platform = detect_platform(url)
                        print(f"[SPY] {platform}: {url[:80]}...")
                        
                        # ===== SHOPIFY: Thử JSON endpoint TRƯỚC (chính xác nhất) =====
                        m = _re.search(r'/products/([^/?#]+)', url)
                        if m and platform not in ('etsy', 'amazon'):
                            handle = m.group(1)
                            from urllib.parse import urlparse as _urlparse
                            parsed = _urlparse(url)
                            json_url = f"{parsed.scheme}://{parsed.netloc}/products/{handle}.json"
                            print(f"  [SHOPIFY] Trying JSON: {json_url}")
                            try:
                                page.goto(json_url, timeout=15000, wait_until="domcontentloaded")
                                time.sleep(1)
                                raw = page.inner_text('body')
                                jdata = json.loads(raw)
                                product = jdata.get("product", {})
                                if product:
                                    from modules.scraper import _clean_title
                                    title = _clean_title(product.get("title", "(No Title)"))
                                    imgs = [img.get("src", "") for img in product.get("images", [])]
                                    imgs = [i for i in imgs if i][:max_img]
                                    results.append({"url": url, "title": title, "images": imgs, "platform": platform})
                                    print(f"  → JSON OK: {title[:50]} | {len(imgs)} images")
                                    continue  # Xong, không cần HTML
                            except Exception as je:
                                print(f"  [SHOPIFY] JSON failed: {je}, fallback to HTML...")
                        
                        # ===== LOAD PAGE (non-Etsy only) =====
                        page.goto(url, timeout=60000, wait_until="domcontentloaded")
                        time.sleep(delay)
                        page.mouse.wheel(0, 3000); time.sleep(1)
                        page.mouse.wheel(0, 3000); time.sleep(1)
                        
                        html = page.content()
                        data = extract_product_data(html, url, max_img)
                        
                        # Log kết quả
                        n_img = len(data.get("images", []))
                        print(f"  → HTML: {data['title'][:50]} | {n_img} images")
                        if n_img == 0:
                            print(f"  [WARN] 0 images! Page length: {len(html)}, platform: {data['platform']}")
                        
                        results.append(data)
                    except Exception as e:
                        print(f"[SPY ERROR] {url}: {traceback.format_exc()}")
                        results.append({"url":url,"title":f"Error: {str(e)[:80]}","images":[],"platform":detect_platform(url),"error":True})
                ctx.close()
                print(f"[CHROME] Done with {len(other_urls)} non-Etsy URLs")
            
            # ===== XỬ LÝ ETSY (persistent profile - login 1 lần) =====
            if etsy_urls:
                print(f"[ETSY] Using persistent profile for {len(etsy_urls)} URLs...")
                
                # Profile riêng cho tool (lưu ở thư mục tool)
                tool_profile = str(APP_DIR / "etsy_profile")
                os.makedirs(tool_profile, exist_ok=True)
                
                try:
                    _cfg2 = load_config()
                    _headless2 = _cfg2.get("headless_scrape", True)
                    ctx_etsy = _launch_persistent(p, tool_profile, headless=_headless2, extra_args=["--no-default-browser-check"])
                    page_etsy = ctx_etsy.pages[0] if ctx_etsy.pages else ctx_etsy.new_page()
                    page_etsy.add_init_script(STEALTH_JS)
                    
                    # Kiểm tra đã login chưa - nếu chưa thì cho user login
                    print(f"[ETSY] Checking session...")
                    try:
                        page_etsy.goto("https://www.etsy.com/", timeout=30000, wait_until="domcontentloaded")
                        time.sleep(3)
                        initial_html = page_etsy.content()
                        
                        if len(initial_html) < 10000:
                            # Bị chặn ngay cả trang home - cần user can thiệp
                            print(f"[ETSY] ⚠ Bị chặn. Cần user login trong cửa sổ Chrome...")
                            print(f"[ETSY] Hãy đăng nhập Etsy trong cửa sổ vừa mở, rồi chờ...")
                            # Chờ user login (tối đa 5 phút)
                            for _ in range(60):
                                time.sleep(5)
                                try:
                                    page_etsy.goto("https://www.etsy.com/", timeout=15000)
                                    time.sleep(2)
                                    if len(page_etsy.content()) > 10000:
                                        print(f"[ETSY] ✓ Login thành công!")
                                        break
                                except:
                                    pass
                    except Exception as e:
                        print(f"[ETSY] Check session error: {e}")
                    
                    # Giờ scrape các URLs
                    for url in etsy_urls:
                        try:
                            print(f"[ETSY] Loading: {url[:80]}...")
                            page_etsy.goto(url, timeout=30000, wait_until="domcontentloaded")
                            time.sleep(3)
                            
                            html = page_etsy.content()
                            if len(html) < 10000:
                                print(f"  ⚠ Page small ({len(html)}), waiting...")
                                time.sleep(5)
                                html = page_etsy.content()
                            
                            if len(html) > 10000:
                                # Scroll lazy images
                                page_etsy.mouse.wheel(0, 2000); time.sleep(1)
                                page_etsy.mouse.wheel(0, 2000); time.sleep(1)
                                html = page_etsy.content()
                                
                                data = extract_product_data(html, url, max_img)
                                results.append(data)
                                print(f"  → {data['title'][:50]} | {len(data.get('images',[]))} images")
                            else:
                                results.append({"url":url,"title":f"Blocked - hãy login lại Etsy","images":[],"platform":"etsy"})
                        except Exception as e:
                            print(f"[ETSY ERROR] {e}")
                            results.append({"url":url,"title":f"Error: {str(e)[:80]}","images":[],"platform":"etsy"})
                    
                    # KHÔNG đóng context để giữ session
                    ctx_etsy.close()
                    print(f"[ETSY] Done, session saved")
                except Exception as e:
                    print(f"[ETSY] Profile error: {e}")
                    import traceback as _tb; _tb.print_exc()
                    for url in etsy_urls:
                        results.append({"url":url,"title":f"Profile error: {str(e)[:80]}","images":[],"platform":"etsy"})
    except Exception as e:
        print(f"[BROWSER ERROR] {traceback.format_exc()}")
        return jsonify({"results":[],"error":str(e)}), 500
    
    return jsonify({"results": results})

# =================== ETSY SHOP JOBS ===================
_shop_jobs = {}  # job_id -> {status, sections, total_products, logs}

def _extract_sections_from_page(page, log):
    """Extract sections từ SIDEBAR trực tiếp — name + count + section_id"""
    return page.evaluate("""
        () => {
            const sections = [];
            const seen = new Set();
            const shopMatch = window.location.pathname.match(/\\/shop\\/([^\\/]+)/);
            const shopName = shopMatch ? shopMatch[1] : '';
            
            // === Strategy 1: Sidebar — tìm mọi element có text "NAME NUMBER" ===
            // Đây là format Etsy sidebar hiển thị
            const candidates = document.querySelectorAll('a, li, button, span, div');
            candidates.forEach(el => {
                const text = (el.textContent || '').trim().replace(/\\s+/g, ' ');
                if (!text || text.length > 100) return;
                
                // Chỉ lấy direct text (không đếm các children lồng nhau)
                // Match "NAME COUNT" ở cuối
                const match = text.match(/^(.{2,60}?)\\s+(\\d{1,5})$/);
                if (!match) return;
                
                const name = match[1].trim();
                const count = parseInt(match[2]);
                const key = name.toLowerCase();
                
                if (seen.has(key)) return;
                
                // === Filter aggressive để loại reviews/announcements/ads ===
                // 1. Count không được là năm (1990-2099)
                if (count >= 1990 && count <= 2099) return;
                // 2. Name không chứa date pattern
                if (/\\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\\b\\s*\\d/i.test(name)) return;
                if (/\\b(january|february|march|april|june|july|august|september|october|november|december)\\b/i.test(name)) return;
                // 3. Name không chứa từ khóa reviews/meta/UI
                if (/\\b(announcement|last updated|on etsy since|sales|already used|collection name|go to slide|review|feedback)/i.test(name)) return;
                // 4. Name không có dấu phẩy ở cuối (dấu hiệu review date)
                if (name.endsWith(',')) return;
                // 5. Name không quá dài (section names thường ngắn)
                if (name.length > 40) return;
                // 6. Loại các từ UI/nav phổ biến
                if (/page|view|more|sort|filter|loading|reviews?|favor|seller|cart|message|sign.in|register|about|polic|follow|share|see all/i.test(name)) return;
                
                if (count < 1 || count > 100000) return;
                
                // Extract section_id từ element hoặc ancestors
                let sid = '';
                // href
                if (el.href) {
                    const m = el.href.match(/section_id=(\\d+)/);
                    if (m) sid = m[1];
                }
                // data-* attributes
                if (!sid) {
                    for (const a of el.attributes || []) {
                        if (/section[-_]?id/i.test(a.name) && /^\\d+$/.test(a.value)) {
                            sid = a.value; break;
                        }
                    }
                }
                // Walk up parents max 5 levels
                if (!sid) {
                    let curr = el.parentElement;
                    for (let i = 0; i < 5 && curr; i++) {
                        if (curr.href) {
                            const m = curr.href.match(/section_id=(\\d+)/);
                            if (m) { sid = m[1]; break; }
                        }
                        for (const a of curr.attributes || []) {
                            if (/section[-_]?id/i.test(a.name) && /^\\d+$/.test(a.value)) {
                                sid = a.value; break;
                            }
                        }
                        if (sid) break;
                        curr = curr.parentElement;
                    }
                }
                
                seen.add(key);
                sections.push({
                    name,
                    url: sid ? `https://www.etsy.com/shop/${shopName}?section_id=${sid}` : '',
                    section_id: sid,
                    count,
                    source: sid ? 'sidebar+id' : 'sidebar'
                });
            });
            
            // === Strategy 2 (fallback): script tags nếu sidebar fail ===
            if (sections.length === 0) {
                document.querySelectorAll('script').forEach(s => {
                    const content = s.textContent || '';
                    if (!content.includes('section_id')) return;
                    const pattern = /"section_id"\\s*:\\s*"?(\\d+)"?[^{}]{0,300}?"(?:title|section_title|name)"\\s*:\\s*"([^"]+)"[^{}]{0,300}?"(?:active_listing_count|listing_count|count)"\\s*:\\s*(\\d+)/g;
                    let m;
                    while ((m = pattern.exec(content)) !== null) {
                        if (seen.has(m[1])) continue;
                        seen.add(m[1]);
                        sections.push({
                            name: m[2].trim(),
                            url: `https://www.etsy.com/shop/${shopName}?section_id=${m[1]}`,
                            section_id: m[1],
                            count: parseInt(m[3]),
                            source: 'script'
                        });
                    }
                });
            }
            
            return sections;
        }
    """)


def _run_sections_job(job_id, shop_url):
    """Bước 1: Quick scan lấy section list"""
    from modules.scraper import STEALTH_JS
    import time, traceback
    from playwright.sync_api import sync_playwright
    
    job = _shop_jobs[job_id]
    def log(m, t='info'):
        job["logs"].append({"m": m, "t": t, "ts": time.strftime('%H:%M:%S')})
    
    try:
        with sync_playwright() as p:
            tool_profile = str(APP_DIR / "etsy_profile")
            os.makedirs(tool_profile, exist_ok=True)
            log("Loading Etsy shop...")
            
            try:
                ctx = _launch_persistent(p, tool_profile, headless=False)
                page = ctx.pages[0] if ctx.pages else ctx.new_page()
                page.add_init_script(STEALTH_JS)
            except Exception as pe:
                log(f"Profile conflict: {pe}", 'err')
                job["status"] = "done"
                return
            
            try:
                # Normalize URL
                import re as _re
                from urllib.parse import urlparse as _up
                parsed = _up(shop_url)
                path_match = _re.match(r'^/shop/([^/]+)', parsed.path)
                if path_match:
                    shop_url = f"https://www.etsy.com/shop/{path_match.group(1)}"
                
                page.goto(shop_url, timeout=60000, wait_until="domcontentloaded")
                time.sleep(4)
                
                for _ in range(4):
                    page.mouse.wheel(0, 1500); time.sleep(0.5)
                page.mouse.wheel(0, -10000); time.sleep(1)
                
                sections_raw = _extract_sections_from_page(page, log)
                log(f"Found {len(sections_raw)} raw sections")
                
                # Filter All/duplicates/latest activity
                sections = []
                seen_names = set()
                for s in sections_raw:
                    nm = s['name'].lower().strip()
                    if nm in ('all', 'all items', 'all products', 'home', 'shop home', 'on sale', 'reviews', 'about', ''): continue
                    # Chỉ bỏ "Latest activity: Apr 15" — KHÔNG bỏ tên section có chứa tháng như "4TH OF JULY", "CINCO DE MAYO"
                    if _re.search(r'^(latest activity|hoạt động)', nm, _re.I): continue
                    # Bỏ format "Apr 15, 2025" etc.
                    if _re.match(r'^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d', nm, _re.I): continue
                    if nm in seen_names: continue
                    seen_names.add(nm)
                    sections.append(s)
                
                job["sections"] = sections
                job["shop_url"] = shop_url
                log(f"✓ {len(sections)} sections (counts from Etsy sidebar)", 'ok')
                ctx.close()
            except Exception as e:
                log(f"Error: {e}", 'err')
                traceback.print_exc()
                try: ctx.close()
                except: pass
    except Exception as e:
        log(f"FATAL: {e}", 'err')
    job["status"] = "done"


def _run_shop_job(job_id, shop_url, selected_sections, delay):
    """Bước 2: Scrape selected sections với thumbnails"""
    from modules.scraper import EtsyExtractor, STEALTH_JS
    import time, traceback
    from playwright.sync_api import sync_playwright
    
    job = _shop_jobs[job_id]
    def log(m, t='info'):
        job["logs"].append({"m": m, "t": t, "ts": time.strftime('%H:%M:%S')})
    
    try:
        with sync_playwright() as p:
            tool_profile = str(APP_DIR / "etsy_profile")
            os.makedirs(tool_profile, exist_ok=True)
            log(f"Scraping {len(selected_sections)} selected sections...")
            
            try:
                ctx = _launch_persistent(p, tool_profile, headless=False)
                page = ctx.pages[0] if ctx.pages else ctx.new_page()
                page.add_init_script(STEALTH_JS)
            except Exception as pe:
                log(f"Profile conflict: {pe}", 'err')
                job["status"] = "done"
                return
            
            try:
                # Load shop home 1 lần
                log("Loading shop home...")
                page.goto(shop_url, timeout=60000, wait_until="domcontentloaded")
                try: page.wait_for_load_state("networkidle", timeout=8000)
                except: pass
                time.sleep(2)
                
                for sec in selected_sections:
                    expected = sec.get('count', 0)
                    sid = sec.get('section_id', '')
                    shop_name_part = shop_url.rstrip('/').split('/')[-1].split('?')[0]
                    
                    log(f"Scanning: {sec['name']} (expect {expected})")
                    products = []
                    
                    # === Nếu không có section_id, click sidebar để lấy ===
                    if not sid:
                        log(f"  no section_id, click sidebar to resolve...")
                        try:
                            # Về shop home
                            page.goto(f"https://www.etsy.com/shop/{shop_name_part}", timeout=30000, wait_until="domcontentloaded")
                            time.sleep(3)
                            # Scroll to load sidebar
                            for _ in range(3):
                                page.mouse.wheel(0, 500); time.sleep(0.3)
                            
                            # Click element có text khớp với tên section
                            url_before = page.url
                            import re as _re2
                            clicked_url = page.evaluate("""
                                async (name) => {
                                    const nameUp = name.trim().toUpperCase();
                                    const candidates = document.querySelectorAll('a, button, [role="link"], [role="button"], span, div');
                                    for (const el of candidates) {
                                        const txt = (el.textContent || '').trim();
                                        if (!txt) continue;
                                        const txtUp = txt.toUpperCase();
                                        // Match: tên section (có thể kèm số) 
                                        const matchExact = txtUp === nameUp;
                                        const matchWithCount = txtUp.startsWith(nameUp + ' ') && /\\d+$/.test(txt);
                                        const matchInLink = txtUp.includes(nameUp) && txt.length < 60;
                                        if (!matchExact && !matchWithCount && !matchInLink) continue;
                                        // Click element
                                        try {
                                            el.click();
                                            await new Promise(r => setTimeout(r, 2500));
                                            // Check URL đã thay đổi sang section_id chưa
                                            if (location.href.includes('section_id=')) {
                                                return location.href;
                                            }
                                        } catch(e) {}
                                    }
                                    return null;
                                }
                            """, sec['name'])
                            
                            if clicked_url:
                                m = _re2.search(r'section_id=(\d+)', clicked_url)
                                if m: 
                                    sid = m.group(1)
                                    log(f"  ✓ resolved section_id={sid} via click")
                        except Exception as ce:
                            log(f"  click error: {ce}", 'warn')
                    
                    # Nếu vẫn không có section_id → skip (KHÔNG dùng search_query vì kết quả sai)
                    if not sid:
                        log(f"  ✗ Cannot resolve section_id. Skipping.", 'err')
                        _update_section_products(job, sec['name'], [], final=True)
                        continue
                    
                    # Navigate đến URL section_id chuẩn
                    target_url = f"https://www.etsy.com/shop/{shop_name_part}?section_id={sid}"
                    log(f"  URL: {target_url}")
                    
                    # Chỉ goto nếu chưa ở URL đúng
                    if f'section_id={sid}' not in page.url:
                        try:
                            page.goto(target_url, timeout=60000, wait_until="domcontentloaded")
                            try: page.wait_for_load_state("networkidle", timeout=10000)
                            except: pass
                            time.sleep(3)  # đợi lâu hơn cho listings render
                        except Exception as ne:
                            log(f"  navigate failed: {ne}", 'err')
                            _update_section_products(job, sec['name'], [], final=True)
                            continue
                    
                    # Verify URL đúng
                    if f'section_id={sid}' not in page.url:
                        log(f"  ⚠ Etsy redirected: {page.url[:100]}", 'warn')
                    
                    # Đợi listings xuất hiện
                    try:
                        page.wait_for_selector('a[href*="/listing/"]', timeout=10000)
                    except:
                        log(f"  no listings found on page", 'err')
                        _update_section_products(job, sec['name'], [], final=True)
                        continue
                    
                    # === MULTI-PAGE: Navigate từng page bằng URL ===
                    all_products_collected = []
                    seen_ids = set()
                    page_num = 1
                    max_pages = max(1, (expected // 24) + 2) if expected > 0 else 5
                    
                    while page_num <= max_pages:
                        # Construct URL cho page hiện tại (giữ section_id!)
                        if page_num == 1:
                            page_url = target_url
                        else:
                            page_url = f"https://www.etsy.com/shop/{shop_name_part}?section_id={sid}&page={page_num}"
                        
                        # Navigate nếu chưa ở đúng URL
                        current = page.url
                        need_navigate = (page_num > 1) or (f'section_id={sid}' not in current)
                        if need_navigate:
                            try:
                                page.goto(page_url, timeout=60000, wait_until="domcontentloaded")
                                try: page.wait_for_load_state("networkidle", timeout=8000)
                                except: pass
                                time.sleep(2)
                            except Exception as ne:
                                log(f"  page {page_num} navigate failed: {ne}", 'err')
                                break
                        
                        # Verify URL vẫn giữ section_id
                        if f'section_id={sid}' not in page.url:
                            log(f"  ⚠ page {page_num}: lost section_id! URL={page.url[:80]}", 'warn')
                            break
                        
                        # Đợi listings
                        try:
                            page.wait_for_selector('a[href*="/listing/"]', timeout=8000)
                        except:
                            log(f"  page {page_num}: no listings found", 'warn')
                            break
                        
                        # Scroll để load lazy
                        stable = 0
                        prev = 0
                        for scroll_i in range(20):
                            page.evaluate("window.scrollBy(0, 2000)")
                            time.sleep(0.5)
                            try: page.wait_for_load_state("networkidle", timeout=2000)
                            except: pass
                            curr = page.evaluate("document.querySelectorAll('a[href*=\"/listing/\"]').length")
                            if curr == prev:
                                stable += 1
                                if stable >= 4: break
                            else:
                                stable = 0; prev = curr
                        
                        # Scroll lại đầu + scroll lại để trigger lazy img
                        page.evaluate("window.scrollTo(0, 0)"); time.sleep(0.5)
                        for _ in range(6):
                            page.evaluate("window.scrollBy(0, 2000)")
                            time.sleep(0.5)
                        
                        # Extract products từ page hiện tại
                        try:
                            page_items = page.evaluate("""
                                (seenIdsArr) => {
                                    const seenIds = new Set(seenIdsArr);
                                    const upgradeImg = (src) => {
                                        if (!src) return '';
                                        return src.replace(/il_\\d+x(?:\\d+|N)\\./g, 'il_570xN.')
                                                  .replace(/iusa_\\d+x\\d+/g, 'iusa_400x400');
                                    };
                                    const bestImg = (img) => {
                                        if (!img) return '';
                                        const srcset = img.getAttribute('srcset') || '';
                                        if (srcset) {
                                            const parts = srcset.split(',').map(s => s.trim().split(/\\s+/));
                                            const sorted = parts.sort((a,b) => parseInt(b[1]||'0') - parseInt(a[1]||'0'));
                                            if (sorted[0]) return sorted[0][0];
                                        }
                                        return upgradeImg(img.currentSrc || img.src || img.getAttribute('data-src') || '');
                                    };
                                    
                                    const results = [];
                                    
                                    // === WHITELIST: Chỉ lấy từ listing cards trong product grid ===
                                    // Cách 1: Tìm cards có data-listing-id (chắc chắn nhất)
                                    const cards = document.querySelectorAll('[data-listing-id], [data-palette-listing-id], [data-listing-card-v2]');
                                    cards.forEach(card => {
                                        const lid = card.getAttribute('data-listing-id') || card.getAttribute('data-palette-listing-id') || '';
                                        const link = card.querySelector('a[href*="/listing/"]');
                                        if (!link) return;
                                        const m = (link.href || '').match(/\\/listing\\/(\\d+)/);
                                        const id = lid || (m ? m[1] : '');
                                        if (!id || seenIds.has(id)) return;
                                        
                                        // Verify card nằm trong product grid (không phải review/sidebar)
                                        const rect = card.getBoundingClientRect();
                                        if (rect.width < 80 || rect.height < 80) return;
                                        
                                        seenIds.add(id);
                                        const img = card.querySelector('img');
                                        const y = rect.top + window.scrollY;
                                        results.push({
                                            url: 'https://www.etsy.com/listing/' + id,
                                            id: id,
                                            thumb: bestImg(img),
                                            y: Math.round(y)
                                        });
                                    });
                                    
                                    // Cách 2 (fallback): Nếu Etsy không có data-listing-id, 
                                    // tìm listing links trong grid chính (có giá + ảnh lớn)
                                    if (results.length === 0) {
                                        const links = document.querySelectorAll('a[href*="/listing/"]');
                                        links.forEach(a => {
                                            const m = (a.href || '').match(/\\/listing\\/(\\d+)/);
                                            if (!m || seenIds.has(m[1])) return;
                                            
                                            // Must have a visible image (product card, not text link in review)
                                            const li = a.closest('li, article, [class*="listing"]');
                                            if (!li) return;
                                            const img = li.querySelector('img');
                                            if (!img) return;
                                            
                                            // Must have price nearby (confirms it's a product card)
                                            const hasPrice = li.textContent && /\\$|USD|\\d+\\.\\d{2}/.test(li.textContent);
                                            if (!hasPrice) return;
                                            
                                            const rect = a.getBoundingClientRect();
                                            if (rect.width < 80 || rect.height < 80) return;
                                            
                                            seenIds.add(m[1]);
                                            const y = rect.top + window.scrollY;
                                            results.push({
                                                url: 'https://www.etsy.com/listing/' + m[1],
                                                id: m[1],
                                                thumb: bestImg(img),
                                                y: Math.round(y)
                                            });
                                        });
                                    }
                                    
                                    results.sort((a, b) => a.y - b.y);
                                    return results;
                                }
                            """, list(seen_ids))
                        except Exception as ex:
                            log(f"  extract error page {page_num}: {ex}", 'err')
                            page_items = []
                        
                        new_count = len(page_items)
                        for item in page_items:
                            seen_ids.add(item['id'])
                        all_products_collected.extend(page_items)
                        log(f"  page {page_num}: {new_count} new items (total {len(all_products_collected)})")
                        
                        # Đủ expected → dừng
                        if expected > 0 and len(all_products_collected) >= expected:
                            break
                        
                        # Không tìm thêm gì → hết products
                        if new_count == 0:
                            log(f"  no new items on page {page_num}, stopping")
                            break
                        
                        page_num += 1
                    
                    # Lấy N đầu tiên nếu có expected
                    if expected > 0:
                        products = all_products_collected[:expected]
                    else:
                        products = all_products_collected
                    products = [{"url": p["url"], "id": p["id"], "thumb": p["thumb"]} for p in products]
                    
                    _update_section_products(job, sec['name'], products, final=True)
                    log(f"✓ {sec['name']}: {len(products)}/{expected} ({page_num} pages)", 'ok' if len(products)>=expected else 'warn')
                    
                
                ctx.close()
            except Exception as e:
                log(f"Error: {e}", 'err')
                traceback.print_exc()
                try: ctx.close()
                except: pass
    except Exception as e:
        log(f"FATAL: {e}", 'err')
    job["status"] = "done"

def _update_section_products(job, name, products, final=False):
    """Update section trong job với products có thumbnail"""
    existing = next((s for s in job["sections"] if s["name"] == name), None)
    if existing:
        existing["products"] = products
        existing["count"] = len(products)
    else:
        job["sections"].append({"name": name, "products": products, "count": len(products)})
    job["total_products"] = sum(s.get("count", 0) for s in job["sections"])

def _update_section(job, name, links, final=False):
    """Legacy: giữ để tương thích"""
    existing = next((s for s in job["sections"] if s["name"] == name), None)
    if existing:
        existing["links"] = links
        existing["count"] = len(links)
    else:
        job["sections"].append({"name": name, "links": links, "count": len(links)})
    job["total_products"] = sum(s["count"] for s in job["sections"])

@flask_app.route("/api/spy/etsy-shop-sections", methods=["POST"])
def etsy_shop_sections_only():
    """Bước 1: Scan shop lấy list sections nhanh"""
    import threading as _th, uuid
    d = request.json
    shop_url = d.get("shop_url", "")
    job_id = uuid.uuid4().hex[:12]
    _shop_jobs[job_id] = {"status": "running", "sections": [], "total_products": 0, "logs": [], "shop_url": shop_url, "phase": "sections"}
    t = _th.Thread(target=_run_sections_job, args=(job_id, shop_url), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})

@flask_app.route("/api/spy/etsy-shop-scrape", methods=["POST"])
def etsy_shop_scrape_selected():
    """Bước 2: Scrape các section đã select"""
    import threading as _th, uuid
    d = request.json
    shop_url = d.get("shop_url", "")
    selected_sections = d.get("sections", [])  # list {name, url, count, section_id}
    delay = d.get("delay", 1.0)
    job_id = uuid.uuid4().hex[:12]
    _shop_jobs[job_id] = {"status": "running", "sections": [], "total_products": 0, "logs": [], "phase": "scrape"}
    t = _th.Thread(target=_run_shop_job, args=(job_id, shop_url, selected_sections, delay), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})

@flask_app.route("/api/spy/etsy-shop-start", methods=["POST"])
def etsy_shop_start():
    """Legacy: vẫn giữ để tương thích"""
    return etsy_shop_sections_only()

@flask_app.route("/api/spy/etsy-shop-status/<job_id>")
def etsy_shop_status(job_id):
    job = _shop_jobs.get(job_id)
    if not job: return jsonify({"error": "Job not found"}), 404
    return jsonify(job)

@flask_app.route("/api/spy/etsy-shop-export", methods=["POST"])
def etsy_shop_export():
    """Export Excel: mỗi section = 1 cột, rows là URL"""
    from openpyxl import Workbook
    import io as _io
    d = request.json
    sections = d.get("sections", [])  # [{name, products|links}]
    
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Shop Products"
    
    # Header row: tên sections
    for ci, sec in enumerate(sections):
        cell = ws.cell(row=1, column=ci+1, value=sec['name'])
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8455EF")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[_col_letter(ci+1)].width = 55
    
    # Rows: URLs
    wrap = Alignment(wrap_text=True, vertical="top")
    max_rows = max((len(s.get('products') or s.get('links') or []) for s in sections), default=0)
    for ci, sec in enumerate(sections):
        items = sec.get('products') or sec.get('links') or []
        for ri, it in enumerate(items):
            url = it if isinstance(it, str) else it.get('url', '')
            cell = ws.cell(row=ri+2, column=ci+1, value=url)
            cell.alignment = wrap
    
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="shop_sections.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _col_letter(n):
    """1 -> A, 2 -> B, ..., 27 -> AA"""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


@flask_app.route("/api/dropbox/test-token", methods=["POST"])
def dbx_test():
    from modules.dropbox_links import test_token
    try: return jsonify(test_token(request.json.get("token","")))
    except Exception as e: return jsonify({"valid":False,"error":str(e)}), 400

@flask_app.route("/api/dropbox/oauth-start", methods=["POST"])
def dbx_oauth_start():
    from modules.dropbox_links import get_oauth_url
    flow, url = get_oauth_url(); flask_app._dbx_flow = flow
    return jsonify({"auth_url": url})

@flask_app.route("/api/dropbox/oauth-finish", methods=["POST"])
def dbx_oauth_finish():
    from modules.dropbox_links import finish_oauth
    token = finish_oauth(flask_app._dbx_flow, request.json.get("code",""))
    cfg = load_config(); cfg["dropbox_token"] = token; save_config(cfg)
    return jsonify({"access_token": token})

_dbx_jobs = {}  # job_id -> {status, products, logs, done, total}

def _run_dbx_job(job_id, access_token, parent_folders):
    """Scan multiple parent folders, mỗi subfolder = 1 product row, streaming"""
    from modules.dropbox_links import scan_subfolders_as_products, natural_key, ensure_shared_link, list_files
    import dropbox, os as _os, time
    
    job = _dbx_jobs[job_id]
    def log(m, t='info'):
        job["logs"].append({"m": m, "t": t, "ts": time.strftime('%H:%M:%S')})
    
    try:
        dbx = dropbox.Dropbox(access_token, timeout=60)
        try:
            acc = dbx.users_get_current_account()
            log(f"✓ Authenticated: {acc.name.display_name}", 'ok')
        except:
            log("Token invalid", 'err')
            job["status"] = "done"
            return
        
        all_products = []
        for fi, parent_path in enumerate(parent_folders):
            p = parent_path.strip().replace("\\", "/")
            if not p.startswith("/"): p = "/" + p
            log(f"[{fi+1}/{len(parent_folders)}] Scan parent: {p}")
            
            try:
                result = dbx.files_list_folder(p, recursive=False)
                subfolders = [e for e in result.entries if isinstance(e, dropbox.files.FolderMetadata)]
                while result.has_more:
                    result = dbx.files_list_folder_continue(result.cursor)
                    subfolders.extend([e for e in result.entries if isinstance(e, dropbox.files.FolderMetadata)])
                
                subfolders.sort(key=lambda x: natural_key(x.name))
                log(f"  Found {len(subfolders)} subfolders")
                job["total"] = (job.get("total", 0) + len(subfolders))
                
                for sf in subfolders:
                    sku = sf.name
                    images = []
                    image_pairs = []
                    try:
                        # Get all files in subfolder
                        result2 = dbx.files_list_folder(sf.path_lower, recursive=False)
                        files = [e for e in result2.entries if isinstance(e, dropbox.files.FileMetadata)]
                        while result2.has_more:
                            result2 = dbx.files_list_folder_continue(result2.cursor)
                            files.extend([e for e in result2.entries if isinstance(e, dropbox.files.FileMetadata)])
                        
                        # Filter ảnh + tạo link SONG SONG (8 API calls cùng lúc)
                        image_files = [f for f in files if _os.path.splitext(f.name)[1].lower() in ('.png','.jpg','.jpeg','.webp','.gif','.bmp')]
                        if image_files:
                            from concurrent.futures import ThreadPoolExecutor, as_completed
                            def _link_one(fm):
                                try:
                                    return (fm.name, ensure_shared_link(dbx, fm.path_lower), None)
                                except Exception as ex:
                                    return (fm.name, None, str(ex))
                            with ThreadPoolExecutor(max_workers=8) as ex:
                                futures = [ex.submit(_link_one, f) for f in image_files]
                                for future in as_completed(futures):
                                    name, url, err = future.result()
                                    if url:
                                        image_pairs.append((name, url))
                                    else:
                                        log(f"    ✗ Link error for {name}: {err}", 'warn')
                        
                        # Sort by natural filename order
                        image_pairs.sort(key=lambda x: natural_key(x[0]))
                        images = [url for _, url in image_pairs]
                        
                        product = {"sku": sku, "images": images}
                        all_products.append(product)
                        job["products"] = all_products
                        job["done"] = job.get("done", 0) + 1
                        log(f"  ✓ {sku}: {len(images)} images", 'ok' if images else 'warn')
                    except Exception as e:
                        log(f"  ✗ {sku}: {e}", 'err')
                        job["done"] = job.get("done", 0) + 1
            except Exception as e:
                log(f"  ✗ Parent folder error: {e}", 'err')
        
        log(f"✓ Done. Total {len(all_products)} products", 'ok')
    except Exception as e:
        log(f"FATAL: {e}", 'err')
        import traceback; traceback.print_exc()
    
    job["status"] = "done"

@flask_app.route("/api/dropbox/scan", methods=["POST"])
def dbx_scan():
    import threading as _th, uuid
    d = request.json
    token = d.get("access_token", "")
    folders = d.get("folders", [])
    if not token: return jsonify({"error": "No token"}), 400
    if not folders: return jsonify({"error": "No folders"}), 400
    
    job_id = uuid.uuid4().hex[:12]
    _dbx_jobs[job_id] = {"status": "running", "products": [], "logs": [], "done": 0, "total": 0}
    t = _th.Thread(target=_run_dbx_job, args=(job_id, token, folders), daemon=True)
    t.start()
    return jsonify({"job_id": job_id})

@flask_app.route("/api/dropbox/status/<job_id>")
def dbx_status(job_id):
    job = _dbx_jobs.get(job_id)
    if not job: return jsonify({"error": "Job not found"}), 404
    return jsonify(job)

@flask_app.route("/api/dropbox/export-excel", methods=["POST"])
def dbx_export_excel():
    """Export Excel: mỗi SKU 1 row, cột A=SKU, B+=links"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io
    d = request.json
    products = d.get("products", [])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dropbox Links"
    
    max_imgs = max((len(p.get('images', [])) for p in products), default=0)
    
    # Header
    headers = ["SKU"] + [f"Image {i+1}" for i in range(max_imgs)]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8455EF")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    wrap = Alignment(wrap_text=True, vertical="top")
    for ri, p in enumerate(products, 2):
        cell = ws.cell(row=ri, column=1, value=p.get('sku', ''))
        cell.alignment = wrap
        for ii, img in enumerate(p.get('images', []), 2):
            cell = ws.cell(row=ri, column=ii, value=img)
            cell.alignment = wrap
    
    # Column widths
    ws.column_dimensions['A'].width = 35
    for i in range(2, max_imgs + 2):
        col_letter = chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)
        ws.column_dimensions[col_letter].width = 55
    
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="dropbox_links.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# =================== LAUNCH ===================
def start_server():
    flask_app.run(host="127.0.0.1", port=5123, debug=False, use_reloader=False)

if __name__ == "__main__":
    url = "http://127.0.0.1:5123"
    
    # === Prevent double instance ===
    import socket
    _check = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        _check.bind(("127.0.0.1", 5123))
        _check.close()
    except OSError:
        # Port đã dùng → app đang chạy rồi, mở browser thay vì start mới
        print("[STARTUP] App already running! Opening browser...")
        webbrowser.open(url)
        sys.exit(0)
    
    print(f"[STARTUP] APP_DIR = {APP_DIR}")
    print(f"[STARTUP] static exists = {_static_dir.exists()}")
    print(f"[STARTUP] frozen = {getattr(sys, 'frozen', False)}")
    
    def find_chrome():
        """Tìm Chrome hoặc Edge để mở App Mode"""
        import platform
        if platform.system() != "Windows":
            return None
        paths = [
            os.path.expandvars(r"%ProgramFiles%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%ProgramFiles(x86)%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%LocalAppData%\Google\Chrome\Application\chrome.exe"),
            os.path.expandvars(r"%ProgramFiles(x86)%\Microsoft\Edge\Application\msedge.exe"),
            os.path.expandvars(r"%ProgramFiles%\Microsoft\Edge\Application\msedge.exe"),
        ]
        for p in paths:
            if os.path.exists(p):
                return p
        return None
    
    def wait_for_server(timeout=15):
        """Đợi server sẵn sàng trước khi mở Chrome"""
        import urllib.request
        for i in range(timeout * 2):
            try:
                urllib.request.urlopen(url, timeout=1)
                print(f"[STARTUP] Server ready after {i*0.5}s")
                return True
            except:
                time.sleep(0.5)
        print("[STARTUP] Server NOT ready after timeout!")
        return False
    
    def open_app_window():
        """Mở app trong cửa sổ riêng + tự tắt server khi đóng"""
        # Đợi server thật sự sẵn sàng
        if not wait_for_server():
            # Server crash → show error
            try:
                import ctypes
                log_path = str(APP_DIR / "app_log.txt")
                ctypes.windll.user32.MessageBoxW(0, 
                    f"Tool Spy Idea không thể khởi động!\n\nXem log tại:\n{log_path}", 
                    "Lỗi", 0x10)
            except:
                pass
            os._exit(1)
        
        chrome = find_chrome()
        if chrome:
            # Dùng profile riêng để Chrome mở process MỚI
            # (nếu dùng Chrome đang mở → merge process → lỗi proxy/extension)
            import tempfile
            app_profile = os.path.join(APP_DIR_BASE, "chrome_app_profile")
            os.makedirs(app_profile, exist_ok=True)
            
            proc = subprocess.Popen([
                chrome,
                f"--app={url}",
                f"--user-data-dir={app_profile}",
                "--window-size=1400,900",
                "--no-first-run",
                "--no-default-browser-check",
                "--disable-extensions",
            ])
            print(f"[STARTUP] Chrome App Mode opened (PID={proc.pid})")
            proc.wait()
            print("[STARTUP] Window closed, shutting down...")
            os._exit(0)
        else:
            webbrowser.open(url)
            print("[STARTUP] Opened in default browser")
    
    print(f"[STARTUP] Starting server on {url}")
    threading.Thread(target=open_app_window, daemon=True).start()
    
    try:
        flask_app.run(host="127.0.0.1", port=5123, debug=False, use_reloader=False)
    except Exception as e:
        print(f"[FATAL] Server crashed: {e}")
        import traceback
        traceback.print_exc()
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(0, f"Server lỗi:\n{e}", "Tool Spy Idea - Error", 0x10)
        except: pass
        os._exit(1)
